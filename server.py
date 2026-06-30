
import os, re, json, time, threading
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl
from werkzeug.utils import secure_filename
from flask import Flask, jsonify, send_from_directory, abort, request
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# ─── Config ────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
DATA_DIR   = BASE_DIR / "data"
JSON_CACHE = BASE_DIR / "data.json"
PUBLIC_DIR = BASE_DIR / "public"
PORT       = 3000

# Simple shared password gate for uploads. Change this, or better, set it
# via an environment variable so it isn't sitting in plain text in the code:
#   export UPLOAD_PASSWORD="something-only-you-know"
UPLOAD_PASSWORD  = os.environ.get("UPLOAD_PASSWORD", "Niyamtek@123")
MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB cap per file

# ─── India City → State dictionary (fallback for unknown cities) ────────────
# This covers major cities across all states
CITY_STATE = {
    # Andhra Pradesh
    "vijayawada":"Andhra Pradesh","visakhapatnam":"Andhra Pradesh","guntur":"Andhra Pradesh",
    "nellore":"Andhra Pradesh","kurnool":"Andhra Pradesh","rajahmundry":"Andhra Pradesh",
    "tirupati":"Andhra Pradesh","kakinada":"Andhra Pradesh","kadapa":"Andhra Pradesh",
    "anantapur":"Andhra Pradesh","ananthapuramu":"Andhra Pradesh","eluru":"Andhra Pradesh",
    "ongole":"Andhra Pradesh","vizianagaram":"Andhra Pradesh","chittoor":"Andhra Pradesh",
    "hindupur":"Andhra Pradesh","bhimavaram":"Andhra Pradesh","machilipatnam":"Andhra Pradesh",
    "adoni":"Andhra Pradesh","tenali":"Andhra Pradesh","proddatur":"Andhra Pradesh",
    "amaravati":"Andhra Pradesh","srikakulam":"Andhra Pradesh","nandyal":"Andhra Pradesh",

    # Arunachal Pradesh
    "itanagar":"Arunachal Pradesh","naharlagun":"Arunachal Pradesh","pasighat":"Arunachal Pradesh",

    # Assam

    "Assam":"Assam","guwahati":"Assam","dibrugarh":"Assam","silchar":"Assam","jorhat":"Assam",
    "nagaon":"Assam","tinsukia":"Assam","tezpur":"Assam","bongaigaon":"Assam",
    "dhubri":"Assam","north lakhimpur":"Assam","karimganj":"Assam","sivasagar":"Assam",

    # Bihar
    "patna":"Bihar","gaya":"Bihar","bhagalpur":"Bihar","muzaffarpur":"Bihar",
    "purnia":"Bihar","darbhanga":"Bihar","bihar sharif":"Bihar","arrah":"Bihar",
    "begusarai":"Bihar","katihar":"Bihar","munger":"Bihar","chhapra":"Bihar",
    "danapur":"Bihar","bettiah":"Bihar","saharsa":"Bihar","sasaram":"Bihar",
    "hajipur":"Bihar","dehri":"Bihar","siwan":"Bihar","motihari":"Bihar",
    "nawada":"Bihar","bagaha":"Bihar","buxar":"Bihar","sitamarhi":"Bihar",
    "kishanganj":"Bihar","aurangabad":"Bihar","jehanabad":"Bihar","arwal":"Bihar",

    # Chandigarh
    "chandigarh":"Chandigarh",

    # Chhattisgarh
    "raipur":"Chhattisgarh","bhilai":"Chhattisgarh","bilaspur":"Chhattisgarh",
    "korba":"Chhattisgarh","durg":"Chhattisgarh","rajnandgaon":"Chhattisgarh",
    "jagdalpur":"Chhattisgarh","raigarh":"Chhattisgarh","ambikapur":"Chhattisgarh",
    "dhamtari":"Chhattisgarh","mahasamund":"Chhattisgarh","baikunthpur":"Chhattisgarh",
    "adawal":"Chhattisgarh",

    # Delhi
    "delhi":"Delhi","new delhi":"Delhi","dwarka":"Delhi","rohini":"Delhi",
    "janakpuri":"Delhi","pitampura":"Delhi","noida":"Delhi",

    # Goa
    "panaji":"Goa","margao":"Goa","vasco da gama":"Goa","mapusa":"Goa",
    "ponda":"Goa","aquem":"Goa","goa":"Goa",

    # Gujarat
    "ahmedabad":"Gujarat","surat":"Gujarat","vadodara":"Gujarat","rajkot":"Gujarat",
    "bhavnagar":"Gujarat","jamnagar":"Gujarat","junagadh":"Gujarat","gandhinagar":"Gujarat",
    "anand":"Gujarat","navsari":"Gujarat","morbi":"Gujarat","nadiad":"Gujarat",
    "surendranagar":"Gujarat","bharuch":"Gujarat","mehsana":"Gujarat","porbandar":"Gujarat",
    "amreli":"Gujarat","valsad":"Gujarat","vapi":"Gujarat","ankleshwar":"Gujarat",
    "godhra":"Gujarat","patan":"Gujarat","botad":"Gujarat","dahod":"Gujarat",
    "kheda":"Gujarat","aravalli":"Gujarat","mahisagar":"Gujarat","chhota udaipur":"Gujarat",
    "tapi":"Gujarat","narmada":"Gujarat","devbhumi dwarka":"Gujarat","gir somnath":"Gujarat",
    "amboli":"Gujarat","anand":"Gujarat","alikherva":"Gujarat","andada":"Gujarat",
    "anklav":"Gujarat","baben":"Gujarat",

    # Haryana
    "faridabad":"Haryana","gurgaon":"Haryana","gurugram":"Haryana","panipat":"Haryana",
    "ambala":"Haryana","yamunanagar":"Haryana","rohtak":"Haryana","hisar":"Haryana",
    "karnal":"Haryana","sonipat":"Haryana","panchkula":"Haryana","bhiwani":"Haryana",
    "sirsa":"Haryana","bahadurgarh":"Haryana","jind":"Haryana","thanesar":"Haryana",
    "kaithal":"Haryana","rewari":"Haryana","palwal":"Haryana","fatehabad":"Haryana",
    "assandh":"Haryana","badhi majra 126":"Haryana","naraingarh":"Haryana",

    # Himachal Pradesh
    "shimla":"Himachal Pradesh","mandi":"Himachal Pradesh","solan":"Himachal Pradesh",
    "dharamsala":"Himachal Pradesh","palampur":"Himachal Pradesh","baddi":"Himachal Pradesh",
    "nahan":"Himachal Pradesh","kullu":"Himachal Pradesh","hamirpur":"Himachal Pradesh",
    "una":"Himachal Pradesh","arki":"Himachal Pradesh","baddi industrial area":"Himachal Pradesh",

    # Jammu & Kashmir
    "srinagar":"Jammu & Kashmir","jammu":"Jammu & Kashmir","anantnag":"Jammu & Kashmir",
    "sopore":"Jammu & Kashmir","baramulla":"Jammu & Kashmir","udhampur":"Jammu & Kashmir",
    "kathua":"Jammu & Kashmir","punch":"Jammu & Kashmir","rajouri":"Jammu & Kashmir",
    "leh":"Jammu & Kashmir","kargil":"Jammu & Kashmir",

    # Jharkhand
    "ranchi":"Jharkhand","jamshedpur":"Jharkhand","dhanbad":"Jharkhand","bokaro":"Jharkhand",
    "deoghar":"Jharkhand","phusro":"Jharkhand","hazaribagh":"Jharkhand","giridih":"Jharkhand",
    "ramgarh":"Jharkhand","medininagar":"Jharkhand","chaibasa":"Jharkhand","chirkunda":"Jharkhand",
    "babupur industrial area":"Jharkhand",

    # Karnataka
    "bengaluru":"Karnataka","bangalore":"Karnataka","mysuru":"Karnataka","mysore":"Karnataka",
    "hubli":"Karnataka","hubballi":"Karnataka","dharwad":"Karnataka","mangaluru":"Karnataka",
    "mangalore":"Karnataka","belagavi":"Karnataka","belgaum":"Karnataka","davangere":"Karnataka",
    "ballari":"Karnataka","bellary":"Karnataka","vijayapura":"Karnataka","bijapur":"Karnataka",
    "shivamogga":"Karnataka","shimoga":"Karnataka","tumkur":"Karnataka","raichur":"Karnataka",
    "bidar":"Karnataka","gulbarga":"Karnataka","kalaburagi":"Karnataka","hassan":"Karnataka",
    "udupi":"Karnataka","chitradurga":"Karnataka","kolar":"Karnataka","mandya":"Karnataka",
    "chikkamagaluru":"Karnataka","bagalkot":"Karnataka","koppal":"Karnataka","gadag":"Karnataka",
    "yadgir":"Karnataka","haveri":"Karnataka","dakshina kannada":"Karnataka",
    "kodagu":"Karnataka","uttara kannada":"Karnataka","chikkaballapur":"Karnataka",
    "ramnagara":"Karnataka","chamarajanagar":"Karnataka","tumkuru":"Karnataka",
    "afzalpur":"Karnataka","aland":"Karnataka","alevoor":"Karnataka","ankola":"Karnataka",
    "arasinakunte":"Karnataka","arsikere":"Karnataka","athni":"Karnataka","aurad":"Karnataka",
    "bada":"Karnataka","badagabettu no 80":"Karnataka","bagepalli":"Karnataka",
    "bail hongal":"Karnataka","belur":"Karnataka","narasimharajapura":"Karnataka",

    # Kerala
    "thiruvananthapuram":"Kerala","kochi":"Kerala","kozhikode":"Kerala","thrissur":"Kerala",
    "kollam":"Kerala","palakkad":"Kerala","alappuzha":"Kerala","malappuram":"Kerala",
    "kannur":"Kerala","kasaragod":"Kerala","kottayam":"Kerala","idukki":"Kerala",
    "ernakulam":"Kerala","wayanad":"Kerala","pathanamthitta":"Kerala",
    "trivandrum":"Kerala","calicut":"Kerala","trichur":"Kerala","quilon":"Kerala",
    "adoor":"Kerala","amballoor":"Kerala","amballur":"Kerala","anchal":"Kerala",
    "ambalaappuzha":"Kerala","ayoor":"Kerala","azhikode":"Kerala","abdurahiman nagar":"Kerala",
    "adat":"Kerala","adichanalloor":"Kerala","adinad":"Kerala","aimanam":"Kerala",
    "akathiyoor":"Kerala","ala":"Kerala","alamcode":"Kerala","alangad":"Kerala",
    "alanthur":"Kerala","alankod":"Kerala","alathur":"Kerala","alur":"Kerala",
    "aluva":"Kerala","ambalapuzha":"Kerala","anakkayam":"Kerala","anthoor":"Kerala",
    "aroor":"Kerala","athirampuzha":"Kerala","athiyannur":"Kerala","atholi":"Kerala",
    "attingal":"Kerala","avittathur":"Kerala","ayancheri":"Kerala",

    # Madhya Pradesh
    "bhopal":"Madhya Pradesh","indore":"Madhya Pradesh","jabalpur":"Madhya Pradesh",
    "gwalior":"Madhya Pradesh","ujjain":"Madhya Pradesh","sagar":"Madhya Pradesh",
    "dewas":"Madhya Pradesh","satna":"Madhya Pradesh","ratlam":"Madhya Pradesh",
    "rewa":"Madhya Pradesh","murwara":"Madhya Pradesh","singrauli":"Madhya Pradesh",
    "burhanpur":"Madhya Pradesh","khandwa":"Madhya Pradesh","bhind":"Madhya Pradesh",
    "chhindwara":"Madhya Pradesh","shivpuri":"Madhya Pradesh","vidisha":"Madhya Pradesh",
    "chhatarpur":"Madhya Pradesh","damoh":"Madhya Pradesh","mandsaur":"Madhya Pradesh",
    "khargone":"Madhya Pradesh","neemuch":"Madhya Pradesh","pithampur":"Madhya Pradesh",
    "narmadapuram":"Madhya Pradesh","hoshangabad":"Madhya Pradesh","itarsi":"Madhya Pradesh",
    "sehore":"Madhya Pradesh","betul":"Madhya Pradesh","seoni":"Madhya Pradesh",
    "datia":"Madhya Pradesh","dhar":"Madhya Pradesh","tikamgarh":"Madhya Pradesh",
    "narsinghpur":"Madhya Pradesh","morena":"Madhya Pradesh","shahdol":"Madhya Pradesh",
    "alirajpur":"Madhya Pradesh","anuppur":"Madhya Pradesh","ashoknagar":"Madhya Pradesh",
    "balaghat":"Madhya Pradesh","barwani":"Madhya Pradesh","ashta":"Madhya Pradesh",
    "badi":"Madhya Pradesh","badnawar":"Madhya Pradesh","badra":"Madhya Pradesh",
    "acharpura industrial area":"Madhya Pradesh","amanganj chhatarpur":"Madhya Pradesh",

    # Maharashtra
    "mumbai":"Maharashtra","pune":"Maharashtra","nagpur":"Maharashtra","thane":"Maharashtra",
    "nashik":"Maharashtra","aurangabad":"Maharashtra","solapur":"Maharashtra",
    "kalyan":"Maharashtra","amravati":"Maharashtra","nanded":"Maharashtra",
    "kolhapur":"Maharashtra","akola":"Maharashtra","latur":"Maharashtra","dhule":"Maharashtra",
    "ahmednagar":"Maharashtra","chandrapur":"Maharashtra","jalgaon":"Maharashtra",
    "parbhani":"Maharashtra","ichalkaranji":"Maharashtra","jalna":"Maharashtra",
    "ambarnath":"Maharashtra","bhiwandi":"Maharashtra","satara":"Maharashtra",
    "sangli":"Maharashtra","navi mumbai":"Maharashtra","vasai":"Maharashtra",
    "virar":"Maharashtra","malegaon":"Maharashtra","ulhasnagar":"Maharashtra",
    "hadapsar":"Maharashtra","badlapur":"Maharashtra","panvel":"Maharashtra",
    "alibag":"Maharashtra","raigad":"Maharashtra","ratnagiri":"Maharashtra",
    "sindhudurg":"Maharashtra","bhandara":"Maharashtra","gadchiroli":"Maharashtra",
    "gondiya":"Maharashtra","hingoli":"Maharashtra","nandurbar":"Maharashtra",
    "osmanabad":"Maharashtra","washim":"Maharashtra","yavatmal":"Maharashtra",
    "ahmadpur":"Maharashtra","ajara":"Maharashtra","akkalkot":"Maharashtra",
    "alandi":"Maharashtra","ambarnath":"Maharashtra","amboli":"Maharashtra",
    "anantpur":"Maharashtra","arvi":"Maharashtra","asangaon":"Maharashtra",
    "ashti gadchiroli":"Maharashtra","auric city":"Maharashtra","aamby valley":"Maharashtra",
    "additional dindori midc":"Maharashtra","additional murbad industrial area":"Maharashtra",
    "additional vinchur midc":"Maharashtra",

    # Manipur
    "imphal":"Manipur","thoubal":"Manipur","bishnupur":"Manipur","churachandpur":"Manipur",
    "senapati":"Manipur","ukhrul":"Manipur","jiribam":"Manipur",

    # Meghalaya
    "shillong":"Meghalaya","tura":"Meghalaya","jowai":"Meghalaya","nongstoin":"Meghalaya",

    # Mizoram
    "aizawl":"Mizoram","lunglei":"Mizoram","saiha":"Mizoram","champhai":"Mizoram",

    # Nagaland
    "kohima":"Nagaland","dimapur":"Nagaland","mokokchung":"Nagaland","wokha":"Nagaland",

    # Odisha
    "bhubaneswar":"Odisha","cuttack":"Odisha","rourkela":"Odisha","brahmapur":"Odisha",
    "berhampur":"Odisha","sambalpur":"Odisha","puri":"Odisha","balasore":"Odisha",
    "bhadrak":"Odisha","baripada":"Odisha","jharsuguda":"Odisha","bargarh":"Odisha",
    "angul":"Odisha","dhenkanal":"Odisha","keonjhar":"Odisha","phulbani":"Odisha",
    "rayagada":"Odisha","koraput":"Odisha","sundargarh":"Odisha","athagad":"Odisha",
    "badmal township":"Odisha",

    # Punjab
    "ludhiana":"Punjab","amritsar":"Punjab","jalandhar":"Punjab","patiala":"Punjab",
    "bathinda":"Punjab","hoshiarpur":"Punjab","mohali":"Punjab","batala":"Punjab",
    "pathankot":"Punjab","moga":"Punjab","abohar":"Punjab","malerkotla":"Punjab",
    "khanna":"Punjab","phagwara":"Punjab","muktsar":"Punjab","barnala":"Punjab",
    "rajpura":"Punjab","firozpur":"Punjab","kapurthala":"Punjab","fazilka":"Punjab",
    "ahmedgarh":"Punjab","anandpur sahib":"Punjab","punjab financial corporation":"Punjab",

    # Rajasthan
    "jaipur":"Rajasthan","jodhpur":"Rajasthan","kota":"Rajasthan","bikaner":"Rajasthan",
    "ajmer":"Rajasthan","udaipur":"Rajasthan","bhilwara":"Rajasthan","alwar":"Rajasthan",
    "bharatpur":"Rajasthan","sikar":"Rajasthan","pali":"Rajasthan","sri ganganagar":"Rajasthan",
    "tonk":"Rajasthan","dausa":"Rajasthan","churu":"Rajasthan","jhunjhunu":"Rajasthan",
    "sriganganagar":"Rajasthan","hanumangarh":"Rajasthan","baran":"Rajasthan",
    "bundi":"Rajasthan","chittorgarh":"Rajasthan","dungarpur":"Rajasthan",
    "jalore":"Rajasthan","jhalawar":"Rajasthan","karauli":"Rajasthan","nagaur":"Rajasthan",
    "pratapgarh":"Rajasthan","rajsamand":"Rajasthan","sawai madhopur":"Rajasthan",
    "sirohi":"Rajasthan","barmer":"Rajasthan","banswara":"Rajasthan","dholpur":"Rajasthan",

    # Sikkim
    "gangtok":"Sikkim","namchi":"Sikkim","gyalshing":"Sikkim","mangan":"Sikkim",

    # Tamil Nadu
    "chennai":"Tamil Nadu","coimbatore":"Tamil Nadu","madurai":"Tamil Nadu",
    "tiruchirappalli":"Tamil Nadu","salem":"Tamil Nadu","tirunelveli":"Tamil Nadu",
    "tiruppur":"Tamil Nadu","vellore":"Tamil Nadu","erode":"Tamil Nadu",
    "thoothukudi":"Tamil Nadu","dindigul":"Tamil Nadu","thanjavur":"Tamil Nadu",
    "ranipet":"Tamil Nadu","sivakasi":"Tamil Nadu","kanchipuram":"Tamil Nadu",
    "udhagamandalam":"Tamil Nadu","hosur":"Tamil Nadu","nagercoil":"Tamil Nadu",
    "kumbakonam":"Tamil Nadu","pudukkottai":"Tamil Nadu","ambur":"Tamil Nadu",
    "nagapattinam":"Tamil Nadu","cuddalore":"Tamil Nadu","virudhunagar":"Tamil Nadu",
    "karur":"Tamil Nadu","dharapuram":"Tamil Nadu","ariyalur":"Tamil Nadu",
    "perambalur":"Tamil Nadu","krishnagiri":"Tamil Nadu","dharmapuri":"Tamil Nadu",
    "namakkal":"Tamil Nadu","villupuram":"Tamil Nadu","kallakurichi":"Tamil Nadu",
    "tiruvarur":"Tamil Nadu","ramanathapuram":"Tamil Nadu","tenkasi":"Tamil Nadu",
    "adaikkakuzhi":"Tamil Nadu","adiramapattinam":"Tamil Nadu","adiyanuthu":"Tamil Nadu",
    "agaram":"Tamil Nadu","alanganallur":"Tamil Nadu","alangayam":"Tamil Nadu",
    "alangayam":"Tamil Nadu","alathur":"Tamil Nadu","amathur":"Tamil Nadu",
    "ambasamudram":"Tamil Nadu","ammapettai":"Tamil Nadu","anaimalai":"Tamil Nadu",
    "anaiyur":"Tamil Nadu","andanallur":"Tamil Nadu","andipatti jakkampatti":"Tamil Nadu",
    "annamalai nagar":"Tamil Nadu","annanji":"Tamil Nadu","annur":"Tamil Nadu",
    "anthiyur":"Tamil Nadu","anupuram dae township":"Tamil Nadu","arakonam":"Tamil Nadu",
    "arani":"Tamil Nadu","aranthangi":"Tamil Nadu","aravakurichi":"Tamil Nadu",
    "aravankad ts":"Tamil Nadu","arcot":"Tamil Nadu","arumanai":"Tamil Nadu",
    "arumbanur":"Tamil Nadu","aruppukkottai":"Tamil Nadu","attur":"Tamil Nadu",

    # Telangana
    "hyderabad":"Telangana","warangal":"Telangana","nizamabad":"Telangana",
    "karimnagar":"Telangana","ramagundam":"Telangana","khammam":"Telangana",
    "mahbubnagar":"Telangana","nalgonda":"Telangana","adilabad":"Telangana",
    "suryapet":"Telangana","miryalaguda":"Telangana","siddipet":"Telangana",
    "narayanpet":"Telangana","bachannapet":"Telangana","secunderabad":"Telangana",
    "sangareddy":"Telangana","mancherial":"Telangana","jagtial":"Telangana",
    "kamareddy":"Telangana","nirmal":"Telangana","vikarabad":"Telangana",
    "wanaparthy":"Telangana","jogulamba gadwal":"Telangana","nagarkurnool":"Telangana",
    "narayanpet":"Telangana","medak":"Telangana","yadadri bhuvanagiri":"Telangana",
    "jangaon":"Telangana","mulugu":"Telangana","bhadradri kothagudem":"Telangana",

    # Tripura
    "agartala":"Tripura","dharmanagar":"Tripura","udaipur":"Tripura","kailasahar":"Tripura",
    "amarpur":"Tripura","ambassa":"Tripura",

    # Uttar Pradesh
    "lucknow":"Uttar Pradesh","kanpur":"Uttar Pradesh","agra":"Uttar Pradesh",
    "varanasi":"Uttar Pradesh","meerut":"Uttar Pradesh","prayagraj":"Uttar Pradesh",
    "allahabad":"Uttar Pradesh","bareilly":"Uttar Pradesh","aligarh":"Uttar Pradesh",
    "moradabad":"Uttar Pradesh","saharanpur":"Uttar Pradesh","gorakhpur":"Uttar Pradesh",
    "noida":"Uttar Pradesh","ghaziabad":"Uttar Pradesh","firozabad":"Uttar Pradesh",
    "jhansi":"Uttar Pradesh","mathura":"Uttar Pradesh","lakhimpur":"Uttar Pradesh",
    "muzaffarnagar":"Uttar Pradesh","rampur":"Uttar Pradesh","shahjahanpur":"Uttar Pradesh",
    "farrukhabad":"Uttar Pradesh","mau":"Uttar Pradesh","hapur":"Uttar Pradesh",
    "etawah":"Uttar Pradesh","mirzapur":"Uttar Pradesh","bulandshahr":"Uttar Pradesh",
    "sambhal":"Uttar Pradesh","amroha":"Uttar Pradesh","hardoi":"Uttar Pradesh",
    "bahraich":"Uttar Pradesh","fatehpur":"Uttar Pradesh","unnao":"Uttar Pradesh",
    "sitapur":"Uttar Pradesh","loni":"Uttar Pradesh","jaunpur":"Uttar Pradesh",
    "sultanpur":"Uttar Pradesh","abupur":"Uttar Pradesh","ajhuwa":"Uttar Pradesh",
    "akbarpur":"Uttar Pradesh","alapur":"Uttar Pradesh","alhaipur":"Uttar Pradesh",
    "amethi":"Uttar Pradesh","aonla":"Uttar Pradesh","ashrafpur kichhauchha":"Uttar Pradesh",
    "aurangabad gadana":"Uttar Pradesh","ayodhya":"Uttar Pradesh","azamgarh":"Uttar Pradesh",
    "baghpat":"Uttar Pradesh","bahjoi":"Uttar Pradesh","auraiya":"Uttar Pradesh",

    # Uttarakhand
    "dehradun":"Uttarakhand","haridwar":"Uttarakhand","roorkee":"Uttarakhand",
    "haldwani":"Uttarakhand","rudrapur":"Uttarakhand","kashipur":"Uttarakhand",
    "rishikesh":"Uttarakhand","pithoragarh":"Uttarakhand","ramnagar":"Uttarakhand",
    "almora":"Uttarakhand","nainital":"Uttarakhand","pauri":"Uttarakhand",
    "tehri":"Uttarakhand","mussoorie":"Uttarakhand","kotdwar":"Uttarakhand",

    # West Bengal
    "kolkata":"West Bengal","asansol":"West Bengal","siliguri":"West Bengal",
    "durgapur":"West Bengal","bardhaman":"West Bengal","burdwan":"West Bengal",
    "malda":"West Bengal","baharampur":"West Bengal","habra":"West Bengal",
    "kharagpur":"West Bengal","bally":"West Bengal","jalpaiguri":"West Bengal",
    "raiganj":"West Bengal","krishnanagar":"West Bengal","haldia":"West Bengal",
    "bankura":"West Bengal","purulia":"West Bengal","midnapore":"West Bengal",
    "adra":"West Bengal","ajodhyanagar":"West Bengal","alipukur":"West Bengal",
    "alipur":"West Bengal","amarshi kasba":"West Bengal","amlagora":"West Bengal",
    "andul":"West Bengal","anup nagar":"West Bengal","argari":"West Bengal",
    "baduria":"West Bengal","bagnan":"West Bengal","bagula":"West Bengal",
    "baksa":"West Bengal","baidyabati":"West Bengal",

    # Dadra & Nagar Haveli
    "silvassa":"Dadra & Nagar Haveli & Daman & Diu","daman":"Dadra & Nagar Haveli & Daman & Diu",
    "diu":"Dadra & Nagar Haveli & Daman & Diu","alok city":"Dadra & Nagar Haveli & Daman & Diu",

    # Puducherry
    "puducherry":"Puducherry","pondicherry":"Puducherry","karaikal":"Puducherry",
    "mahe":"Puducherry","yanam":"Puducherry",

    # Andaman & Nicobar Islands
    "port blair":"Andaman & Nicobar Islands",

    # Additional spellings / districts / towns found missing in real data
    "thane district":"Maharashtra","palghar":"Maharashtra","wardha":"Maharashtra",
    "howrah":"West Bengal",
    "davanagere":"Karnataka","kalburgi":"Karnataka",
    "tirupur":"Tamil Nadu","tiruchirapalli":"Tamil Nadu","kanyakumari":"Tamil Nadu",
    "thirupathur":"Tamil Nadu","trichy":"Tamil Nadu","tuticorin":"Tamil Nadu",
    "tiruvallur":"Tamil Nadu","thiruvallur":"Tamil Nadu",
    "bicholim":"Goa","corlim":"Goa","guirim":"Goa","merces":"Goa","benaulim":"Goa",
    "sancoale":"Goa","salvador do mundo":"Goa","curchorem":"Goa",
    "gandhidham-kutch":"Gujarat","gandhidham":"Gujarat",
    "muktsar sahib":"Punjab","mansa":"Punjab","gurdaspur":"Punjab",
    "fatehgarh sahib":"Punjab",
    "mahabubnagar":"Telangana",
    "faizabad":"Uttar Pradesh","raebareli":"Uttar Pradesh",
    "gautam buddh nagar":"Uttar Pradesh","greater noida":"Uttar Pradesh",
    "bijnor":"Uttar Pradesh",
    "chandragiri":"Andhra Pradesh","tirupathi":"Andhra Pradesh",
    "guntakal":"Andhra Pradesh","gudiwada":"Andhra Pradesh","chittor":"Andhra Pradesh",
    "assam":"Assam",
    "sonepat":"Haryana","yamuna nagar":"Haryana","kurukshetra":"Haryana",
    "kangra":"Himachal Pradesh",
    "khurda":"Odisha","jajpur":"Odisha","pipili":"Odisha",
    "vaishali":"Bihar",

    # Batch 2 — high frequency remaining gaps
    "chhatrapati sambhajinagar":"Maharashtra","maraimalai nagar":"Tamil Nadu",
    "matheran":"Maharashtra","karjat":"Maharashtra","robertsonpet":"Karnataka",
    "koregaon bhima":"Maharashtra","puttur":"Karnataka","khapoli":"Maharashtra",
    "khopoli":"Maharashtra","rajamahendravaram":"Andhra Pradesh","chalthan":"Gujarat",
    "eklahare":"Maharashtra","barasat":"West Bengal","ranasan industrial area":"Gujarat",
    "pimpri chinchwad":"Maharashtra","kelambakkam":"Tamil Nadu","mira bhayandar":"Maharashtra",
    "harij":"Gujarat","shirur":"Maharashtra","samalkot":"Andhra Pradesh","karad":"Maharashtra",
    "sayan":"Gujarat","punalur":"Kerala","kottarakkara":"Kerala","ozar":"Maharashtra",
    "etah":"Uttar Pradesh","chakan":"Maharashtra","berhampore":"West Bengal",
    "ghazipur":"Uttar Pradesh","bansberia":"West Bengal","neyyattinkara":"Kerala",
    "boisar":"Maharashtra","karakulam":"Kerala","baraut":"Uttar Pradesh","raybag":"Karnataka",
    "ferozepur":"Punjab","malayinkeezhu":"Kerala","bhuj":"Gujarat","dapoli":"Maharashtra",
    "gopavaram":"Andhra Pradesh","contai":"West Bengal","saundatti yellamma":"Karnataka",
    "banthla":"West Bengal","faridkot":"Punjab","mainpuri":"Uttar Pradesh",
    "palsana":"Gujarat","vyara":"Gujarat","chiplun":"Maharashtra","olpad":"Gujarat",
    "shirdi":"Maharashtra","parangipettai":"Tamil Nadu","tadepalligudem":"Andhra Pradesh",
    "pudussery":"Kerala","panihati":"West Bengal","padrauna":"Uttar Pradesh",
    "kudal":"Maharashtra","kancheepuram":"Tamil Nadu","basti":"Uttar Pradesh",
    "nazira":"Assam","diamond harbour":"West Bengal","rajpur sonarpur":"West Bengal",
    "narasapur":"Andhra Pradesh","kathlal":"Gujarat","jaynagar mazilpur":"West Bengal",
    "kakdwip":"West Bengal","chalakudy":"Kerala","kanhangad":"Kerala","thodupuzha":"Kerala",
}


def get_state(city_raw):
    """Look up state from city name."""
    if not city_raw:
        return ""
    city_clean = str(city_raw).strip()

    # If city already contains ", State" format (from Baanknet location)
    parts = [p.strip() for p in city_clean.split(",")]
    if len(parts) >= 2:
        state_candidate = parts[-1]
        # Validate it looks like a state name (not a PIN or number)
        if state_candidate and not state_candidate.isdigit() and len(state_candidate) > 3:
            return state_candidate

    # Look up in dictionary (case-insensitive)
    key = city_clean.lower().strip()
    if key in CITY_STATE:
        return CITY_STATE[key]

    # Try matching just the first word/part before comma
    first_part = parts[0].lower().strip()
    if first_part in CITY_STATE:
        return CITY_STATE[first_part]

    # Try stripping common suffixes like "District", "(Dist)", etc.
    stripped = re.sub(r"\s*\(?\b(district|dist|taluka|tehsil)\b\)?\s*$", "", key).strip()
    if stripped and stripped != key and stripped in CITY_STATE:
        return CITY_STATE[stripped]

    # Try fuzzy contains-match: see if any known city name is a substring of
    # this city, or vice versa (helps catch minor spelling variants)
    for known_city, known_state in CITY_STATE.items():
        if len(known_city) > 4 and (known_city in key or key in known_city):
            return known_state

    return ""


# ─── In-memory cache ───────────────────────────────────────────────────────
cached_data  = []
last_updated = None


def to_date_str(val):
    if not val or val == "":
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if re.match(r"^\d{5}$", s):
        try:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(s))).strftime("%Y-%m-%d")
        except:
            return s
    m = re.match(r"(\d{2})-(\d{2})-(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    if re.match(r"\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    return s


def clean(val):
    if val is None:
        return ""
    if isinstance(val, float) and val != val:
        return ""
    return str(val).strip()


def find_url(row, headers, candidates):
    """Look for a URL value across several possible column header names."""
    for name in candidates:
        if name in headers:
            try:
                val = clean(row[headers.index(name)])
                if val:
                    return val
            except:
                pass
    return ""


URL_COLUMN_CANDIDATES = [
    "detail_url", "property_link", "url", "listing_url", "source_url", "link",
    "property_url", "auction_link", "page_link", "web_link", "view_link",
]


# ─── Parsers ───────────────────────────────────────────────────────────────
def parse_baanknet(ws):
    rows_iter = ws.iter_rows(values_only=True)
    headers   = [clean(h).lower() for h in next(rows_iter)]

    def col(row, name):
        try: return row[headers.index(name)]
        except: return ""

    records = []
    for row in rows_iter:
        location = clean(col(row, "location"))
        parts    = [p.strip() for p in location.split(",")]
        city     = parts[0] if parts else ""
        state    = get_state(location)  # uses full "city, state" string

        records.append({
            "ID"          : clean(col(row, "property_id")) or clean(col(row, "id")),
            "Bank"        : clean(col(row, "bank")),
            "PropertyType": "",
            "City"        : city,
            "State"       : state,
            "Date"        : to_date_str(col(row, "auction_start_date")),
            "Price"       : clean(col(row, "reserve_price")),
            "closed_date" : to_date_str(col(row, "closed_date")),
            "status"      : clean(col(row, "property_status")),
            "source"      : find_url(row, headers, URL_COLUMN_CANDIDATES),
            "portal"      : "Baanknet",
        })
    return records


def parse_foreclosure(ws):
    rows_iter = ws.iter_rows(values_only=True)
    headers   = [clean(h).lower() for h in next(rows_iter)]

    def col(row, name):
        try: return row[headers.index(name)]
        except: return ""

    records = []
    for row in rows_iter:
        city  = clean(col(row, "city"))
        state = get_state(city)

        records.append({
            "ID"          : clean(col(row, "listing_id")),
            "Bank"        : clean(col(row, "institution_name")),
            "PropertyType": clean(col(row, "property_details")),
            "City"        : city,
            "State"       : state,
            "Date"        : to_date_str(col(row, "auction_date")),
            "Price"       : clean(col(row, "reserve_price")),
            "closed_date" : to_date_str(col(row, "closed_date")),
            "status"      : clean(col(row, "status")),
            "source"      : find_url(row, headers, URL_COLUMN_CANDIDATES),
            "portal"      : "Foreclosure India",
        })
    return records


def detect_and_parse(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        first_row = [clean(h).lower() for h in next(rows_iter)]
        wb.close()

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        if "property_id" in first_row or ("bank" in first_row and "location" in first_row):
            print(f"  → Baanknet format")
            result = parse_baanknet(ws)
        elif "institution_name" in first_row:
            print(f"  → Foreclosure India format")
            result = parse_foreclosure(ws)
        else:
            print(f"  ⚠️  Unknown format: {first_row}")
            result = []

        wb.close()
        return result
    except Exception as e:
        print(f"  ❌ Error: {e}")
        return []


# ─── Load all Excel files ──────────────────────────────────────────────────
def load_all_files():
    global cached_data, last_updated

    DATA_DIR.mkdir(exist_ok=True)
    xlsx_files = list(DATA_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print("⚠️  No .xlsx files found in data/ folder")
        cached_data = []
        return

    all_records = []
    for f in xlsx_files:
        print(f"📄 Reading: {f.name}")
        time.sleep(0.3)
        records = detect_and_parse(f)
        records = [r for r in records if r.get("ID") or r.get("Bank")]
        print(f"     {len(records)} rows loaded")
        all_records.extend(records)

    # Deduplicate
    seen, unique = set(), []
    for r in all_records:
        key = r["ID"] + r.get("portal", "")
        if key not in seen:
            seen.add(key)
            unique.append(r)

    cached_data  = unique
    last_updated = datetime.now().isoformat()
    JSON_CACHE.write_text(json.dumps(unique, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"✅ Total: {len(unique)} unique records from {len(xlsx_files)} file(s)  [{datetime.now().strftime('%H:%M:%S')}]")


def load_data():
    global cached_data, last_updated
    try:
        load_all_files()
    except Exception as e:
        print(f"❌ Load error: {e}")
        if JSON_CACHE.exists():
            cached_data  = json.loads(JSON_CACHE.read_text(encoding="utf-8"))
            last_updated = datetime.fromtimestamp(JSON_CACHE.stat().st_mtime).isoformat()
            print(f"ℹ️  Loaded {len(cached_data)} rows from cache")


# ─── File Watcher ──────────────────────────────────────────────────────────
class ExcelChangeHandler(FileSystemEventHandler):
    def __init__(self):
        self._timer = None

    def on_modified(self, event):
        if event.src_path.endswith(".xlsx"):
            self._debounce()

    def on_created(self, event):
        if event.src_path.endswith(".xlsx"):
            self._debounce()

    def _debounce(self):
        if self._timer:
            self._timer.cancel()
        self._timer = threading.Timer(1.5, lambda: (print("\n🔄 Excel changed — reloading..."), load_data()))
        self._timer.start()


def start_watcher():
    DATA_DIR.mkdir(exist_ok=True)
    handler  = ExcelChangeHandler()
    observer = Observer()
    observer.schedule(handler, str(DATA_DIR), recursive=False)
    observer.start()
    print(f"👀 Watching: {DATA_DIR}")
    return observer


# ─── Flask App ─────────────────────────────────────────────────────────────
app = Flask(__name__, static_folder=str(PUBLIC_DIR))

@app.route("/api/data")
def api_data():
    return jsonify({"ok": True, "count": len(cached_data), "lastUpdated": last_updated, "data": cached_data})

@app.route("/api/reload")
def api_reload():
    load_data()
    return jsonify({"ok": True, "count": len(cached_data), "message": "Reloaded"})

@app.route("/api/upload", methods=["POST"])
def api_upload():
    # Password check — sent as a form field named "password"
    if request.form.get("password") != UPLOAD_PASSWORD:
        return jsonify({"ok": False, "error": "Incorrect upload password"}), 401

    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No file in request"}), 400

    f = request.files["file"]
    if not f.filename or not f.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "Only .xlsx files are accepted"}), 400

    # Guard against oversized uploads
    f.seek(0, os.SEEK_END)
    size = f.tell()
    f.seek(0)
    if size > MAX_UPLOAD_BYTES:
        return jsonify({"ok": False, "error": "File too large (max 10 MB)"}), 400

    DATA_DIR.mkdir(exist_ok=True)

    # Keep the name safe and unique so two people uploading "auctions.xlsx"
    # don't overwrite each other
    safe_name = secure_filename(f.filename) or "upload.xlsx"
    stem, ext = os.path.splitext(safe_name)
    final_name = safe_name
    counter = 1
    while (DATA_DIR / final_name).exists():
        final_name = f"{stem}_{counter}{ext}"
        counter += 1

    save_path = DATA_DIR / final_name
    f.save(str(save_path))

    # Validate it's actually a readable Excel file before keeping it
    try:
        wb = openpyxl.load_workbook(str(save_path), read_only=True)
        wb.close()
    except Exception:
        save_path.unlink(missing_ok=True)
        return jsonify({"ok": False, "error": "That file isn't a valid Excel (.xlsx) file"}), 400

    load_data()  # the watcher would catch this too, but reload immediately for instant feedback
    return jsonify({"ok": True, "filename": final_name, "count": len(cached_data), "message": "Uploaded and merged"})

@app.route("/api/status")
def api_status():
    files = [f.name for f in DATA_DIR.glob("*.xlsx")] if DATA_DIR.exists() else []
    return jsonify({"ok": True, "files": files, "rows": len(cached_data), "lastUpdated": last_updated})

@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_frontend(path):
    target = PUBLIC_DIR / path
    if path and target.exists() and target.is_file():
        return send_from_directory(str(PUBLIC_DIR), path)
    return send_from_directory(str(PUBLIC_DIR), "index.html")

# ─── Start ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("\n" + "="*55)
    print("  🏦  Property Auction Dashboard — Python Backend")
    print("="*55)
    load_data()
    observer = start_watcher()
    print(f"\n🚀 Dashboard: http://localhost:{PORT}\nPress Ctrl+C to stop.\n")
    try:
        app.run(host="0.0.0.0", port=PORT, debug=False, use_reloader=False)
    except KeyboardInterrupt:
        print("\n👋 Shutting down...")
    finally:
        observer.stop()
        observer.join()
